from colorama import init
from colorama import Fore
init()
print(Fore.GREEN + "Rapid7 Report Formatting" + Fore.RESET)

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import openpyxl.utils.dataframe
import itertools
from itertools import islice
import threading
import time
import sys
import daemon
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import re
import collections
from os import getlogin

def poc1():
    print("Please select the csv file to be formatted.")
    root = tk.Tk()
    root.withdraw()
    f = filedialog.askopenfilename()
    uname = getlogin()


    df1 = pd.read_csv(f)
    df1.drop_duplicates(subset=None, inplace=True)
    dfdays = df1['Vulnerability Age'].replace({'Days':''}, regex=True)
    df1.update(dfdays)
    dfdays = df1['Vulnerability Age'].replace({'Day':''}, regex=True)
    df1.update(dfdays)
    df1.update(df1['Vulnerability Age'].astype(int))
    df1.update(pd.to_numeric(df1['Asset OS Version'], errors = 'coerce'))

    pdf = df1['Vulnerability CVSSv3 Score']

    newf = pd.ExcelWriter(fr'C:\Users\{uname}\Desktop\20XX-XX-XX Client Name Rapid7 Vulnerability Report.xlsx')
    # newf = pd.ExcelWriter()
    df1.to_excel(newf, index=False)
    newf.save()

    wb = load_workbook(fr'C:\Users\{uname}\Desktop\20XX-XX-XX Client Name Rapid7 Vulnerability Report.xlsx')
    ws = wb.active
    ws.title = "Report"

    ca = Alignment(horizontal='left',vertical='center')
    ra = Alignment(horizontal='right',vertical='center')
    ws.column_dimensions['A'].width = 15
    ws['A1'].alignment = ca
    ws.column_dimensions['B'].width = 13
    ws['B1'].alignment = ca
    ws.column_dimensions['C'].width = 13
    ws['C1'].alignment = ca
    ws.column_dimensions['D'].width = 20
    ws['D1'].alignment = ca
    ws.column_dimensions['E'].width = 15

    ws['E1'].alignment = ca
    ws.column_dimensions['F'].width = 20
    ws['F1'].alignment = ca
    ws.column_dimensions['G'].width = 20
    ws['G1'].alignment = ca
    ws.column_dimensions['H'].width = 15
    ws['H1'].alignment = ca
    ws.column_dimensions['I'].width = 23
    ws['I1'].alignment = ca
    ws['I1'].value = 'Vulnerability Age (in days)'
    ws.column_dimensions['J'].width = 15
    ws['J1'].alignment = ca
    ws.column_dimensions['K'].width = 20
    ws['K1'].alignment = ca
    ws.column_dimensions['L'].width = 15
    ws['L1'].alignment = ca
    ws.column_dimensions['M'].width = 17
    ws['M1'].alignment = ca
    ws.column_dimensions['N'].width = 10
    ws['N1'].alignment = ca
    ws.column_dimensions['O'].width = 17
    ws['O1'].alignment = ca
    ws.column_dimensions['P'].width = 13
    ws['P1'].alignment = ca
    ws.column_dimensions['Q'].width = 11
    ws['Q1'].alignment = ca
    ws.column_dimensions['R'].width = 13
    ws['R1'].alignment = ca
    ws.column_dimensions['S'].width = 12
    ws['S1'].alignment = ca
    ws.column_dimensions['T'].width = 17
    ws['T1'].alignment = ca
    ws.column_dimensions['U'].width = 18
    ws['U1'].alignment = ca


    ws.insert_cols(14)
    thin = Side(border_style='thin')
    ws['N1'].value = 'Priority'
    ws['N1'].font = Font(bold=True)
    ws['N1'].border = Border(left=thin,right=thin, bottom=thin)
    cn = ws['N']

    for row, data in enumerate(pdf, start=2):
        if data == 0:
            ws.cell(row,14,'None')
        elif data >= 0.1 and data <= 3.9:
            ws.cell(row, 14, 'Low')
        elif data >= 4.0 and data <= 6.9:
            ws.cell(row, 14, 'Medium')
        elif data >= 7.0 and data <= 8.9:
            ws.cell(row, 14, 'High')
        else:
            ws.cell(row, 14, 'Critical')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:V1'
    wb.save(fr'C:\Users\{uname}\Desktop\20XX-XX-XX Client Name Rapid7 Vulnerability Report.xlsx')



    book = newf

    def cvsscount():

        #Sets parameters for animation
        done = False
        def count():
            for c in itertools.cycle(['|','/','-', '\\']):
                if done:
                    print('\n')
                    break
                sys.stdout.write('\rWorking...' + c)
                sys.stdout.flush()
                time.sleep(0.1)
            sys.stdout.write('Converted to excel format! The report has been saved to your desktop. See the summary info below.')
        a = threading.Thread(target=count)
        a.daemon=True
        a.start()

    #Declares IP address pattern for matching, sets counters, opens Excel document and counts number of each CVSSv3 Score
        pattern = re.compile(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})')
        count30Day = 0
        countIP = 0
        count7zip = 0
        countAdobe = 0
        countApache = 0
        countArtifex = 0
        countAtlassian = 0
        countCIFS = 0
        countDefender = 0
        countFirefox = 0
        countGoogle = 0
        countIBM = 0
        countJava = 0
        countMicrosoft = 0
        countMongo = 0
        countOffice = 0
        countPHP = 0
        countRedHat = 0
        countSSH = 0
        countSolaris = 0
        countSymantec = 0
        countTLS = 0
        countWinRAR = 0
        countWireshark = 0
        countVMWare = 0

    #Creates data frames to gather values for summarizing
        dfscore = pd.read_excel(book, 'Report'.capitalize())['Vulnerability CVSSv3 Score']

        dfdays = pd.read_excel(book, 'Report'.capitalize())['Vulnerability Age (in days)']

        dfboth = pd.concat([dfdays,dfscore], axis=1,sort = False)
        dfboth['Vulnerability CVSSv3 Score'] = dfboth['Vulnerability CVSSv3 Score'].astype(str).astype(float)

        iplst = pd.read_excel(book, 'Report'.capitalize())['Asset IP Address'].tolist()
        iplst = list(dict.fromkeys(iplst))
        for line in iplst:
            line = line.strip()
            match = pattern.search(line)
            if match:
                countIP += 1

        cvss = dfscore.value_counts().sort_index()
        days = dfdays.value_counts().sort_index()

        hc30 = (dfboth['Vulnerability Age (in days)'] > 30) & (dfboth['Vulnerability CVSSv3 Score'] > 6.9)
        for rows in hc30:
             if rows == True:
                 count30Day = count30Day + 1
    #Counts number of each vulnerability for top 4 category

        dftitle = pd.read_excel(book, 'Report'.capitalize())['Vulnerability Title']
        titlelst = dftitle.tolist()
        countlst = []
        for line in titlelst:
            line = line.strip()
            if line.find('7-Zip') != -1:
                count7zip += 1
                countlst.append('7-Zip')
            if line.find('Adobe') != -1:
                countAdobe += 1
                countlst.append('Adobe')
            if line.find('Apache') != -1:
                countApache += 1
                countlst.append('Apache')
            if line.find('Artifex') != -1:
                countArtifex += 1
                countlst.append('Artifex')
            if line.find('Atlassian') != -1:
                countAtlassian += 1
                countlst.append('Atlassian')
            if line.find('CIFS') != -1:
                countCIFS += 1
                countlst.append('CIFS')
            if line.find('Defender') != -1:
                countDefender += 1
                countlst.append('Defender')
            if line.find('Firefox') != -1:
                countFirefox += 1
                countlst.append('Firefox')
            if line.find('Google') != -1:
                countGoogle += 1
                countlst.append('Google')
            if line.find('IBM') != -1:
                countIBM += 1
                countlst.append('IBM')
            if line.find('Java') != -1:
                countJava += 1
                countlst.append('Java')
            if line.find('Microsoft') != -1:
                countMicrosoft += 1
            if line.find('Mongo') != -1:
                countMongo += 1
            if line.find('Office') != -1:
                countOffice += 1
            if line.find('PHP') != -1:
                countPHP += 1
                countlst.append('PHP')
            if line.find('Red Hat') != -1:
                countRedHat += 1
                countlst.append('Red Hat')
            if line.find('SSH') != -1:
                countSSH += 1
                countlst.append('SSH')
            if line.find('Solaris') != -1:
                countSolaris += 1
                countlst.append('Solaris')
            if line.find('Symantec') != -1:
                countSymantec += 1
                countlst.append('Symantec')
            if line.find('TLS') != -1:
                countTLS += 1
                countlst.append('TLS')
            if line.find('WinRAR') != -1:
                countWinRAR += 1
                countlst.append('WinRAR')
            if line.find('Wireshark') != -1:
                countWireshark += 1
                countlst.append('Wireshark')
            if line.find('VMWare') != -1:
                countVMWare += 1
                countlst.append('VMWare')

        done = True
        time.sleep(0.5)
        stop_threads = True

        counts = {'7-Zip:': count7zip, 'Adobe:': countAdobe, 'Apache:': countApache, 'Artifex:': countArtifex, 'Atlassian:': countAtlassian, 'CIFS:': countCIFS, 'Defender:': countDefender,
        'Firefox:': countFirefox, 'Google:': countGoogle, 'IBM:':countIBM, 'PHP:': countPHP, 'SSH:': countSSH, 'Solaris: ': countSolaris, 'Symantec:':countSymantec, 'TLS:':countTLS, 'WinRAR:':countWinRAR,
        'Wireshark:':countWireshark, 'VMWare:':countVMWare, 'Java:':countJava, 'Microsoft:':countMicrosoft,'Red Hat:':countRedHat, 'Mongo:':countMongo, 'Office:':countOffice}
        sort_counts = sorted(counts.items(), key=lambda x: x[1], reverse = True)
    #Outputs data into CLI for copying onto summary page
        print('\n' + '\n' + 'Asset Count:',countIP)
        print('\n' + 'Priority (CVSSv3 Score):')
        print("\rNone (0):",cvss.loc[0.0])
        print("\rLow (0.1 - 3.9):",cvss.loc[0.1:3.9].sum())
        print("\rMedium (4.0 - 6.9):",cvss.loc[4.0:6.9].sum())
        print(Fore.RED + "\rHigh (7.0 - 8.9):",cvss.loc[7.0:8.9].sum())
        print("\rCritical (9.0 - 10.0):",cvss.loc[9.0:10.0].sum())

        print('\n'+"\rHigh & Critical >30 Days:",count30Day)
        print(Fore.RESET +"\rTotal Vulns:",cvss.loc[0.0:10.0].sum())

        print('\n' + "\rTop 5 Vulnerabilities")
        for i in sort_counts[0:5]:
            print(i[0], i[1])

    cvsscount()

try:

    poc1()

    done = True
    time.sleep(0.5)
    sys.stdout.write(Fore.GREEN + '\n'  + 'Done!' + '\n' + Fore.RESET)

except:
     print("An unexpected error occured, please try again!")

input("Press Enter to exit.")
